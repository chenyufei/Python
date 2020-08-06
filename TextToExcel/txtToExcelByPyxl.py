# coding=utf-8
import os
from openpyxl import Workbook

os.chdir('E:/pycharm/Test/profile/')
file_path = 'all.xlsx'  # 要写入的文件
wb = Workbook()


def write_excel(txt_name, sheet):
    f1 = open(txt_name, 'r', encoding='utf-8')
    lines = f1.readlines()
    for line in lines:
        split_line = line.split('\t')
        sheet.append(split_line)


app_list = ['brand', 'cate1', 'cate2', 'app_name', 'uv']
attribute_list = ['category', 'brand', 'keyword', 'uv']
city_list = ['category', 'brand', 'city_num', 'cate2']
keyword_list = ['brand', 'keyword', 'uv']
profile_list = ['category', 'brand', 'cate1', 'cate2', 'uv/pv']
profile_cate1_list = ['category', 'brand', 'cate1', 'uv/pv']
star_list = ['brand', 'appid', 'star_name', 'uv']
join_uv_list = ['category', 'brand', 'uv']

sheet = wb.create_sheet('app')
sheet.append(app_list)
write_excel('app.txt', sheet)

sheet = wb.create_sheet('attribute')
sheet.append(attribute_list)
write_excel('attribute.txt', sheet)

sheet = wb.create_sheet('city')
sheet.append(city_list)
write_excel('city.txt', sheet)

sheet = wb.create_sheet('keyword')
sheet.append(keyword_list)
write_excel('keyword.txt', sheet)

sheet = wb.create_sheet('profile')
sheet.append(profile_list)
write_excel('profile.txt', sheet)

sheet = wb.create_sheet('profile_cate1')
sheet.append(profile_cate1_list)
write_excel('profile.txt', sheet)

sheet = wb.create_sheet('star')
sheet.append(star_list)
write_excel('star.txt', sheet)

sheet = wb.create_sheet('join_uv')
sheet.append(join_uv_list)
write_excel('join_uv.txt', sheet)

wb.save(file_path)